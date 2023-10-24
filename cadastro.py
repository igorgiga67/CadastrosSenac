from flask import Flask, render_template, request, redirect, url_for, flash
import pandas as pd
from flask_mail import Mail, Message
import openpyxl

app = Flask(__name__)
app.secret_key = 'some_secret_key'

cadastros = {}
max_pessoas_por_horario = 20
arquivo_excel = "reservas.xlsx"

app.config['MAIL_SERVER'] = 'smtp.office365.com'  # Substitua 'your-email-provider.com' pelo servidor SMTP do seu provedor de e-mail.
app.config['MAIL_PORT'] = 587  # Porta para envio de e-mails. 587 é comum para conexões não seguras, 465 para seguras.
app.config['MAIL_USERNAME'] = 'igor-67@hotmail.com'  # Seu endereço de e-mail
app.config['MAIL_PASSWORD'] = 'urvucrzdeegrkxuy'  # Sua senha de e-mail
app.config['MAIL_USE_TLS'] = True  # Use isso para conexões não seguras
app.config['MAIL_USE_SSL'] = False  # Se estiver usando a porta 465, defina isso como True

mail = Mail(app)

@app.route('/')
def index():
    return render_template('index.html', cadastros=cadastros)

@app.route('/reservar', methods=['POST'])
def reservar():
    horario_reserva = request.form.get('horario')
    nome = request.form.get('nome')
    email = request.form.get('email')
    empresa = request.form.get('empresa')
    telefone = request.form.get('telefone')
    
    # Verificar campos obrigatórios
    if not all([nome, email, empresa, telefone, horario_reserva]):
        flash('Todos os campos são obrigatórios!')
        return redirect(url_for('index'))

    # Tentar ler o arquivo Excel, se não existir, criar um novo dataframe
    try:
        df = pd.read_excel(arquivo_excel)
        new_id = df["ID"].max() + 1 if "ID" in df.columns and not df["ID"].empty else 1
    except FileNotFoundError:
        df = pd.DataFrame(columns=['ID', 'horario', 'nome', 'email', 'empresa', 'telefone'])
        new_id = 1

    # Adicionar nova reserva ao dataframe
    new_entry = {
        'ID': [new_id],  # As entradas precisam ser listas
        'horario': [horario_reserva],
        'nome': [nome],
        'email': [email],
        'empresa': [empresa],
        'telefone': [telefone]
    }
 
    # Adicionar nova entrada ao DataFrame existente
    new_entry_df = pd.DataFrame(new_entry)
    df = pd.concat([df, new_entry_df], ignore_index=True)

 
    # Salvar o DataFrame
    df.to_excel(arquivo_excel, index=False)
    
    # Recarregar as reservas na memória
    load_reservas_from_excel()

    # Feedback ao usuário
    flash(f"Reserva feita para {nome} às {horario_reserva}!")
    return redirect(url_for('index'))


def load_reservas_from_excel():
    global cadastros
    cadastros.clear()
    try:
        df = pd.read_excel(arquivo_excel)
        for _, row in df.iterrows():
            horario = row['horario']
            reserva = {
                'ID': row['ID'],
                'nome': row['nome'],
                'email': row['email'],
                'empresa': row['empresa'],
                'telefone': row['telefone']
            }

            if horario not in cadastros:
                cadastros[horario] = []
            
            cadastros[horario].append(reserva)
    except FileNotFoundError:
        print("Arquivo de reservas não encontrado. Começando com uma lista vazia.")

@app.route('/salvar')
def salvar():
    data = []

    for horario, reservas in cadastros.items():
        for reserva in reservas:
            reserva_data = [
                reserva['ID'],
                horario,
                reserva['nome'],
                reserva['email'],
                reserva['empresa'],
                reserva['telefone']
            ]
            data.append(reserva_data)


    # Crie o DataFrame usando a ordem exata das colunas
    df = pd.DataFrame(data, columns=['ID', 'horario', 'nome', 'email', 'empresa', 'telefone'])
    df.to_excel(arquivo_excel, index=False)
    flash(f"Reservas salvas em {arquivo_excel}!")
    
    msg = Message('Reservas Salvas', sender='igor-67@hotmail.com', recipients=['igorgiga67@gmail.com'])
    msg.body = 'Olá, tudo bem? Meu nome é Igor, e sou Instrutor no Senac MT. Este email é um email automático referente a RQ - 060 que foi preenchida no evento que está ocorrendo neste momento.'
    
    caminho_arquivo = r'C:\Users\igor-\OneDrive\Área de Trabalho\cadastro\CadastrosSenac\reservas.xlsx'
    
    with open(caminho_arquivo, 'rb') as fp:
        msg.attach("reservas.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", fp.read())

    mail.send(msg)

    flash(f"Reservas salvas em {arquivo_excel} e e-mail enviado!")
    return redirect(url_for('index'))

@app.route('/excluir/<int:reserva_id>', methods=['GET'])
def excluir_reserva(reserva_id):
    reservas = []
    
    # Carregue as reservas do arquivo Excel
    with open(arquivo_excel, "rb") as f:
        workbook = openpyxl.load_workbook(f)
        sheet = workbook.active
        
        # Crie um header para o novo arquivo, ainda não sabemos se precisaremos dele
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Se o ID da linha for o que queremos excluir, ignore esta linha
            if row[0] == reserva_id:
                continue
                
            reserva = {
                'ID': row[0],
                'horario': row[1],
                'nome': row[2],
                'email': row[3],
                'empresa': row[4],
                'telefone': row[5]
            }
            reservas.append(reserva)

    # Sobrescreva o arquivo Excel sem a reserva excluída
    novo_workbook = openpyxl.Workbook()
    novo_sheet = novo_workbook.active
    
    novo_sheet.append(headers)  # Adicione o header de volta
    for reserva in reservas:
        novo_sheet.append([reserva['ID'], reserva['horario'], reserva['nome'], reserva['email'], reserva['empresa'], reserva['telefone']])
    novo_workbook.save(arquivo_excel)
    
    cadastros.clear()  # Limpe o dicionário cadastros

    # Recarregue as reservas do arquivo Excel atualizado
    df = pd.read_excel(arquivo_excel)
    for _, row in df.iterrows():
        horario = row['horario']
        reserva = {
            'ID': row['ID'],
            'nome': row['nome'],
            'email': row['email'],
            'empresa': row['empresa'],
            'telefone': row['telefone']
        }

        if horario not in cadastros:
            cadastros[horario] = []

        cadastros[horario].append(reserva)

    flash('Reserva excluída com sucesso!')
    return redirect(url_for('index'))

if __name__ == '__main__':
 
    try:
        df = pd.read_excel(arquivo_excel)
        for _, row in df.iterrows():
            horario = row['horario']
            reserva = {
                'ID' : row['ID'],
                'nome': row['nome'],
                'email': row['email'],
                'empresa': row['empresa'],
                'telefone': row['telefone']
            }

            if horario not in cadastros:
                cadastros[horario] = []

            cadastros[horario].append(reserva)
    except FileNotFoundError:
        print("Arquivo de reservas não encontrado. Começando com uma lista vazia.")

    app.run(debug=True)
